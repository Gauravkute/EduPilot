import openpyxl

def extract(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath,data_only=True)
    sheets = {}

    for name in wb.sheetnames:
        ws = wb[name]
        rows = []

        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        sheets[name] = rows

    return {"sheet_count":len(wb.sheetnames),
            "sheets":sheets}